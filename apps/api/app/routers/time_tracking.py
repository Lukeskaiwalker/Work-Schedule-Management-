from __future__ import annotations
import calendar
import csv
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO, StringIO
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.db import get_db
from app.core.deps import get_current_user
from app.core.permissions import has_permission_for_user
from app.core.time import utcnow
from app.models.entities import BreakEntry, ClockEntry, SchoolAbsence, User, VacationRequest
from app.schemas.api import (
    RequiredDailyHoursOut,
    RequiredDailyHoursUpdate,
    SchoolAbsenceCreate,
    SchoolAbsenceOut,
    TimeCurrentOut,
    TimeEntryOut,
    TimeEntryUpdate,
    TimesheetOut,
    VacationRequestCreate,
    VacationRequestOut,
    VacationRequestReview,
)

router = APIRouter(prefix="/time", tags=["time-tracking"])

# ── Predefined German absence types ───────────────────────────────────────────
ABSENCE_TYPES: list[dict] = [
    {"key": "vacation",       "label_de": "Urlaub",                "label_en": "Vacation",          "counts_as_hours": True},
    {"key": "sick",           "label_de": "Krankheit",             "label_en": "Sick leave",         "counts_as_hours": True},
    {"key": "school",         "label_de": "Berufsschule",          "label_en": "Vocational school",  "counts_as_hours": True},
    {"key": "holiday",        "label_de": "Feiertag",              "label_en": "Public holiday",     "counts_as_hours": True},
    {"key": "special_leave",  "label_de": "Sonderurlaub",          "label_en": "Special leave",      "counts_as_hours": True},
    {"key": "training",       "label_de": "Fortbildung",           "label_en": "Training",           "counts_as_hours": True},
    {"key": "work_accident",  "label_de": "Arbeitsunfall",         "label_en": "Work accident",      "counts_as_hours": True},
    {"key": "company_event",  "label_de": "Betriebsveranstaltung", "label_en": "Company event",      "counts_as_hours": True},
    {"key": "unpaid_leave",   "label_de": "Unbezahlter Urlaub",    "label_en": "Unpaid leave",       "counts_as_hours": False},
    {"key": "parental_leave", "label_de": "Elternzeit",            "label_en": "Parental leave",     "counts_as_hours": False},
    {"key": "care_leave",     "label_de": "Pflegezeit",            "label_en": "Care leave",         "counts_as_hours": False},
    {"key": "other",          "label_de": "Sonstige",              "label_en": "Other",              "counts_as_hours": True},
]

_ABSENCE_TYPE_MAP: dict[str, dict] = {a["key"]: a for a in ABSENCE_TYPES}
_WEEKDAY_ABBR_DE = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]


def _week_bounds(day: date) -> tuple[date, date]:
    start = day - timedelta(days=day.weekday())
    end = start + timedelta(days=6)
    return start, end


def _app_timezone() -> ZoneInfo:
    timezone_name = (get_settings().app_timezone or "UTC").strip() or "UTC"
    try:
        return ZoneInfo(timezone_name)
    except ZoneInfoNotFoundError:
        return ZoneInfo("UTC")


def _local_date_from_utc(utc_naive: datetime) -> date:
    return utc_naive.replace(tzinfo=timezone.utc).astimezone(_app_timezone()).date()


def _local_period_bounds_utc(start_day: date, end_day: date) -> tuple[datetime, datetime]:
    app_tz = _app_timezone()
    start_local = datetime.combine(start_day, time.min, tzinfo=app_tz)
    end_local = datetime.combine(end_day, time.max, tzinfo=app_tz)
    start_utc = start_local.astimezone(timezone.utc).replace(tzinfo=None)
    end_utc = end_local.astimezone(timezone.utc).replace(tzinfo=None)
    return start_utc, end_utc


def _hours_between(start: datetime, end: datetime | None) -> float:
    if not end:
        return 0.0
    return max((end - start).total_seconds(), 0) / 3600.0


def _required_break_hours(worked_hours: float) -> float:
    # German ArbZG baseline: >6h => 30m, >9h => 45m.
    if worked_hours > 9:
        return 0.75
    if worked_hours > 6:
        return 0.5
    return 0.0


def _break_hours(db: Session, clock_entry_id: int, now: datetime | None = None) -> float:
    breaks = db.scalars(select(BreakEntry).where(BreakEntry.clock_entry_id == clock_entry_id)).all()
    total = 0.0
    for br in breaks:
        total += _hours_between(br.break_start, br.break_end or now)
    return total


def _overlap_hours(start: datetime, end: datetime | None, range_start: datetime, range_end: datetime) -> float:
    if not end:
        return 0.0
    overlap_start = max(start, range_start)
    overlap_end = min(end, range_end)
    if overlap_end <= overlap_start:
        return 0.0
    return (overlap_end - overlap_start).total_seconds() / 3600.0


def _break_hours_for_period(
    db: Session,
    clock_entry_id: int,
    period_start: datetime,
    period_end: datetime,
    now: datetime | None = None,
) -> float:
    breaks = db.scalars(select(BreakEntry).where(BreakEntry.clock_entry_id == clock_entry_id)).all()
    total = 0.0
    for br in breaks:
        total += _overlap_hours(br.break_start, br.break_end or now, period_start, period_end)
    return total


def _entry_metrics_for_period(
    db: Session,
    entry: ClockEntry,
    period_start: datetime,
    period_end: datetime,
    now: datetime | None = None,
) -> dict[str, float]:
    end = entry.clock_out or now
    worked_hours = _overlap_hours(entry.clock_in, end, period_start, period_end)
    break_hours = _break_hours_for_period(db, entry.id, period_start, period_end, now=now)
    required_break_hours = _required_break_hours(worked_hours)
    deducted_break_hours = max(break_hours, required_break_hours)
    return {
        "worked_hours": round(worked_hours, 2),
        "break_hours": round(break_hours, 2),
        "required_break_hours": round(required_break_hours, 2),
        "deducted_break_hours": round(deducted_break_hours, 2),
        "net_hours": round(max(worked_hours - deducted_break_hours, 0), 2),
    }


def _entries_overlapping_period(
    db: Session,
    user_id: int,
    period_start: datetime,
    period_end: datetime,
) -> list[ClockEntry]:
    return list(
        db.scalars(
            select(ClockEntry).where(
                ClockEntry.user_id == user_id,
                ClockEntry.clock_in <= period_end,
                or_(ClockEntry.clock_out.is_(None), ClockEntry.clock_out >= period_start),
            )
        ).all()
    )


def _is_time_manager(user: User) -> bool:
    return has_permission_for_user(user.id, user.role, "time:view_all") or \
           has_permission_for_user(user.id, user.role, "time:manage")


def _is_required_hours_manager(user: User) -> bool:
    return has_permission_for_user(user.id, user.role, "time:manage")


def _is_vacation_reviewer(user: User) -> bool:
    return has_permission_for_user(user.id, user.role, "time:approve_vacation")


def _is_school_manager(user: User) -> bool:
    return has_permission_for_user(user.id, user.role, "time:manage_absences")


def _resolve_target_user_id(current_user: User, user_id: int | None) -> int:
    if user_id is None or user_id == current_user.id:
        return current_user.id
    if not _is_time_manager(current_user):
        raise HTTPException(status_code=403, detail="Not allowed")
    return user_id


def _vacation_request_out(db: Session, request_row: VacationRequest) -> VacationRequestOut:
    request_user = db.get(User, request_row.user_id)
    return VacationRequestOut(
        id=request_row.id,
        user_id=request_row.user_id,
        user_name=request_user.display_name if request_user else f"#{request_row.user_id}",
        start_date=request_row.start_date,
        end_date=request_row.end_date,
        note=request_row.note,
        status=request_row.status,
        reviewed_by=request_row.reviewed_by,
        reviewed_at=request_row.reviewed_at,
        created_at=request_row.created_at,
    )


def _school_absence_out(db: Session, row: SchoolAbsence) -> SchoolAbsenceOut:
    target_user = db.get(User, row.user_id)
    return SchoolAbsenceOut(
        id=row.id,
        user_id=row.user_id,
        user_name=target_user.display_name if target_user else f"#{row.user_id}",
        title=row.title,
        absence_type=row.absence_type,
        counts_as_hours=row.counts_as_hours,
        start_date=row.start_date,
        end_date=row.end_date,
        recurrence_weekday=row.recurrence_weekday,
        recurrence_until=row.recurrence_until,
        created_by=row.created_by,
        created_at=row.created_at,
    )


def _get_open_entry(db: Session, user_id: int) -> ClockEntry | None:
    return db.scalars(
        select(ClockEntry).where(ClockEntry.user_id == user_id, ClockEntry.clock_out.is_(None)).order_by(ClockEntry.id.desc())
    ).first()


def _entry_metrics(db: Session, entry: ClockEntry, now: datetime | None = None) -> dict[str, float]:
    end = entry.clock_out or now
    worked_hours = _hours_between(entry.clock_in, end)
    break_hours = _break_hours(db, entry.id, now=now)
    required_break_hours = _required_break_hours(worked_hours)
    deducted_break_hours = max(break_hours, required_break_hours)
    return {
        "worked_hours": round(worked_hours, 2),
        "break_hours": round(break_hours, 2),
        "required_break_hours": round(required_break_hours, 2),
        "deducted_break_hours": round(deducted_break_hours, 2),
        "net_hours": round(max(worked_hours - deducted_break_hours, 0), 2),
    }


def _entry_out(db: Session, entry: ClockEntry, now: datetime | None = None) -> TimeEntryOut:
    metrics = _entry_metrics(db, entry, now=now)
    return TimeEntryOut(
        id=entry.id,
        user_id=entry.user_id,
        clock_in=entry.clock_in,
        clock_out=entry.clock_out,
        is_open=entry.clock_out is None,
        break_hours=metrics["break_hours"],
        required_break_hours=metrics["required_break_hours"],
        deducted_break_hours=metrics["deducted_break_hours"],
        net_hours=metrics["net_hours"],
    )


def _sanitized_required_daily_hours(user: User) -> float:
    return round(max(float(user.required_daily_hours or 8.0), 1.0), 2)


@router.post("/clock-in")
def clock_in(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    open_entry = _get_open_entry(db, current_user.id)
    if open_entry:
        raise HTTPException(status_code=400, detail="Already clocked in")

    entry = ClockEntry(user_id=current_user.id)
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return {"clock_entry_id": entry.id, "clock_in": entry.clock_in}


@router.post("/clock-out")
def clock_out(
    clock_entry_id: int | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    entry = db.get(ClockEntry, clock_entry_id) if clock_entry_id else _get_open_entry(db, current_user.id)
    if not entry or entry.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Clock entry not found or no open shift")
    if entry.clock_out is not None:
        raise HTTPException(status_code=400, detail="Already clocked out")

    now = utcnow()
    open_break = db.scalars(
        select(BreakEntry).where(BreakEntry.clock_entry_id == entry.id, BreakEntry.break_end.is_(None))
    ).first()
    if open_break:
        open_break.break_end = now

    entry.clock_out = now
    db.add(entry)
    db.commit()
    db.refresh(entry)
    metrics = _entry_metrics(db, entry)
    return {
        "clock_entry_id": entry.id,
        "clock_out": entry.clock_out,
        "worked_hours": metrics["worked_hours"],
        "break_hours": metrics["break_hours"],
        "required_break_hours": metrics["required_break_hours"],
        "deducted_break_hours": metrics["deducted_break_hours"],
        "net_hours": metrics["net_hours"],
    }


@router.post("/{clock_entry_id}/break-start")
def break_start(clock_entry_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    entry = db.get(ClockEntry, clock_entry_id)
    if not entry or entry.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Clock entry not found")
    if entry.clock_out is not None:
        raise HTTPException(status_code=400, detail="Shift already closed")

    existing = db.scalars(
        select(BreakEntry).where(BreakEntry.clock_entry_id == clock_entry_id, BreakEntry.break_end.is_(None))
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Break already open")

    br = BreakEntry(clock_entry_id=clock_entry_id)
    db.add(br)
    db.commit()
    db.refresh(br)
    return {"break_id": br.id, "break_start": br.break_start}


@router.post("/break-start")
def break_start_open(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    entry = _get_open_entry(db, current_user.id)
    if not entry:
        raise HTTPException(status_code=400, detail="No open shift")
    return break_start(entry.id, current_user=current_user, db=db)


@router.post("/{clock_entry_id}/break-end")
def break_end(clock_entry_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    entry = db.get(ClockEntry, clock_entry_id)
    if not entry or entry.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Clock entry not found")

    br = db.scalars(
        select(BreakEntry).where(BreakEntry.clock_entry_id == clock_entry_id, BreakEntry.break_end.is_(None))
    ).first()
    if not br:
        raise HTTPException(status_code=400, detail="No open break")

    br.break_end = utcnow()
    db.add(br)
    db.commit()
    db.refresh(br)
    return {"break_id": br.id, "break_end": br.break_end}


@router.post("/break-end")
def break_end_open(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    entry = _get_open_entry(db, current_user.id)
    if not entry:
        raise HTTPException(status_code=400, detail="No open shift")
    return break_end(entry.id, current_user=current_user, db=db)


@router.get("/current", response_model=TimeCurrentOut)
def current_status(
    user_id: int | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    target_user_id = _resolve_target_user_id(current_user, user_id)
    target_user = db.get(User, target_user_id)
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    now = utcnow()
    required_daily_hours = _sanitized_required_daily_hours(target_user)

    local_today = _local_date_from_utc(now)
    start_dt, end_dt = _local_period_bounds_utc(local_today, local_today)
    day_entries = _entries_overlapping_period(db, target_user_id, start_dt, end_dt)
    daily_net_hours = round(
        sum(_entry_metrics_for_period(db, day_entry, start_dt, end_dt, now=now)["net_hours"] for day_entry in day_entries),
        2,
    )
    progress_percent = round((daily_net_hours / required_daily_hours) * 100, 2) if required_daily_hours > 0 else 0.0

    entry = _get_open_entry(db, target_user_id)
    if not entry:
        return TimeCurrentOut(
            server_time=now,
            required_daily_hours=required_daily_hours,
            daily_net_hours=daily_net_hours,
            progress_percent_live=progress_percent,
        )

    metrics = _entry_metrics(db, entry, now=now)
    open_break = db.scalars(
        select(BreakEntry).where(BreakEntry.clock_entry_id == entry.id, BreakEntry.break_end.is_(None))
    ).first()
    return TimeCurrentOut(
        server_time=now,
        clock_entry_id=entry.id,
        clock_in=entry.clock_in,
        break_open=open_break is not None,
        worked_hours_live=metrics["worked_hours"],
        break_hours_live=metrics["break_hours"],
        required_break_hours_live=metrics["required_break_hours"],
        deducted_break_hours_live=metrics["deducted_break_hours"],
        net_hours_live=metrics["net_hours"],
        required_daily_hours=required_daily_hours,
        daily_net_hours=daily_net_hours,
        progress_percent_live=progress_percent,
    )


@router.patch("/required-hours/{user_id}", response_model=RequiredDailyHoursOut)
def set_required_daily_hours(
    user_id: int,
    payload: RequiredDailyHoursUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not _is_required_hours_manager(current_user):
        raise HTTPException(status_code=403, detail="Not allowed")

    target_user = db.get(User, user_id)
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    target_user.required_daily_hours = round(payload.required_daily_hours, 2)
    db.add(target_user)
    db.commit()
    db.refresh(target_user)
    return RequiredDailyHoursOut(
        user_id=target_user.id,
        required_daily_hours=target_user.required_daily_hours,
    )


@router.post("/vacation-requests", response_model=VacationRequestOut)
def create_vacation_request(
    payload: VacationRequestCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if payload.end_date < payload.start_date:
        raise HTTPException(status_code=400, detail="end_date must be on or after start_date")
    row = VacationRequest(
        user_id=current_user.id,
        start_date=payload.start_date,
        end_date=payload.end_date,
        note=(payload.note or "").strip() or None,
        status="pending",
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _vacation_request_out(db, row)


@router.get("/vacation-requests", response_model=list[VacationRequestOut])
def list_vacation_requests(
    status: str | None = None,
    user_id: int | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    target_user_id = user_id
    if not _is_time_manager(current_user):
        target_user_id = current_user.id
    stmt = select(VacationRequest)
    if status:
        stmt = stmt.where(VacationRequest.status == status.strip().lower())
    if target_user_id:
        stmt = stmt.where(VacationRequest.user_id == target_user_id)
    rows = db.scalars(stmt.order_by(VacationRequest.created_at.desc(), VacationRequest.id.desc())).all()
    return [_vacation_request_out(db, row) for row in rows]


@router.patch("/vacation-requests/{request_id}", response_model=VacationRequestOut)
def review_vacation_request(
    request_id: int,
    payload: VacationRequestReview,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not _is_vacation_reviewer(current_user):
        raise HTTPException(status_code=403, detail="Not allowed")
    row = db.get(VacationRequest, request_id)
    if not row:
        raise HTTPException(status_code=404, detail="Vacation request not found")
    next_status = payload.status.strip().lower()
    if next_status not in {"approved", "rejected"}:
        raise HTTPException(status_code=400, detail="Invalid status")
    row.status = next_status
    row.reviewed_by = current_user.id
    row.reviewed_at = utcnow()
    db.add(row)
    db.commit()
    db.refresh(row)
    return _vacation_request_out(db, row)


@router.get("/school-absences", response_model=list[SchoolAbsenceOut])
def list_school_absences(
    user_id: int | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    target_user_id = user_id
    if not _is_time_manager(current_user):
        target_user_id = current_user.id
    stmt = select(SchoolAbsence)
    if target_user_id:
        stmt = stmt.where(SchoolAbsence.user_id == target_user_id)
    rows = db.scalars(stmt.order_by(SchoolAbsence.start_date.desc(), SchoolAbsence.id.desc())).all()
    return [_school_absence_out(db, row) for row in rows]


@router.post("/school-absences", response_model=SchoolAbsenceOut)
def create_school_absence(
    payload: SchoolAbsenceCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not _is_school_manager(current_user):
        raise HTTPException(status_code=403, detail="Not allowed")
    target_user = db.get(User, payload.user_id)
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    if payload.end_date < payload.start_date:
        raise HTTPException(status_code=400, detail="end_date must be on or after start_date")
    if payload.recurrence_weekday is not None and payload.recurrence_until and payload.recurrence_until < payload.start_date:
        raise HTTPException(status_code=400, detail="recurrence_until must be on or after start_date")
    if payload.recurrence_weekday is not None and payload.end_date != payload.start_date:
        raise HTTPException(
            status_code=400,
            detail="For recurring school days use a single-day start/end date and recurrence_until",
        )
    # Derive counts_as_hours from the canonical absence type if not overridden by client
    type_meta = _ABSENCE_TYPE_MAP.get(payload.absence_type, {})
    counts = payload.counts_as_hours if "counts_as_hours" in payload.model_fields_set else type_meta.get("counts_as_hours", True)

    row = SchoolAbsence(
        user_id=payload.user_id,
        title=payload.title.strip(),
        absence_type=payload.absence_type,
        counts_as_hours=counts,
        start_date=payload.start_date,
        end_date=payload.end_date,
        recurrence_weekday=payload.recurrence_weekday,
        recurrence_until=payload.recurrence_until,
        created_by=current_user.id,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _school_absence_out(db, row)


@router.delete("/school-absences/{absence_id}")
def delete_school_absence(
    absence_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not _is_school_manager(current_user):
        raise HTTPException(status_code=403, detail="Not allowed")
    row = db.get(SchoolAbsence, absence_id)
    if not row:
        raise HTTPException(status_code=404, detail="School absence not found")
    db.delete(row)
    db.commit()
    return {"ok": True}


def _nrw_public_holidays(year: int) -> dict[date, str]:
    """Return NRW public holidays for the given year as {date: German name}.

    Uses the Anonymous Gregorian Easter algorithm. All 13 NRW statutory
    holidays are included (9 fixed + 4 moveable/variable).
    """
    # Easter Sunday via Anonymous Gregorian algorithm
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    easter = date(year, month, day)

    return {
        date(year, 1, 1):  "Neujahrstag",
        easter - timedelta(days=2): "Karfreitag",
        easter:                      "Ostersonntag",
        easter + timedelta(days=1):  "Ostermontag",
        date(year, 5, 1):  "Tag der Arbeit",
        easter + timedelta(days=39): "Christi Himmelfahrt",
        easter + timedelta(days=49): "Pfingstsonntag",
        easter + timedelta(days=50): "Pfingstmontag",
        easter + timedelta(days=60): "Fronleichnam",
        date(year, 10, 3): "Tag der Deutschen Einheit",
        date(year, 11, 1): "Allerheiligen",
        date(year, 12, 25): "1. Weihnachtstag",
        date(year, 12, 26): "2. Weihnachtstag",
    }


def _credited_absence_hours_for_period(
    db: Session,
    user_id: int,
    start_date: date,
    end_date: date,
    required_daily_hours: float,
    clocked_dates: set[date],
) -> float:
    """Return hours credited from public holidays, approved vacation, and school absences.

    Only weekdays (Mon–Fri) are considered. Days where the employee already has a
    clock entry are skipped (no double-counting). NRW public holidays that fall on
    weekdays are credited automatically without any DB record needed.
    Days with counts_as_hours=False absences contribute 0.
    """
    credited = 0.0
    # Pre-build holiday sets for all years touched by the period
    holiday_sets: dict[int, dict[date, str]] = {}
    for yr in range(start_date.year, end_date.year + 1):
        holiday_sets[yr] = _nrw_public_holidays(yr)

    cursor = start_date
    while cursor <= end_date:
        if cursor.weekday() < 5 and cursor not in clocked_dates:
            # 1. Public holiday automatically credits hours
            if cursor in holiday_sets.get(cursor.year, {}):
                credited += required_daily_hours
                cursor += timedelta(days=1)
                continue
            # 2. Check school/other absences
            absence = db.scalars(
                select(SchoolAbsence).where(
                    SchoolAbsence.user_id == user_id,
                    SchoolAbsence.start_date <= cursor,
                    SchoolAbsence.end_date >= cursor,
                )
            ).first()
            if absence:
                if absence.counts_as_hours:
                    credited += required_daily_hours
            else:
                # 3. Check approved vacation
                vacation = db.scalars(
                    select(VacationRequest).where(
                        VacationRequest.user_id == user_id,
                        VacationRequest.status == "approved",
                        VacationRequest.start_date <= cursor,
                        VacationRequest.end_date >= cursor,
                    )
                ).first()
                if vacation:
                    credited += required_daily_hours
        cursor += timedelta(days=1)
    return round(credited, 2)


@router.get("/absence-types")
def get_absence_types():
    return ABSENCE_TYPES


@router.get("/public-holidays")
def get_public_holidays(
    year: int | None = None,
    current_user: User = Depends(get_current_user),
):
    """Return NRW public holidays for a given year (defaults to current year).

    Response: list of {date: "YYYY-MM-DD", name: "..."}
    """
    target_year = year or _local_date_from_utc(utcnow()).year
    holidays = _nrw_public_holidays(target_year)
    return [
        {"date": d.isoformat(), "name": name}
        for d, name in sorted(holidays.items())
    ]


@router.get("/timesheet", response_model=TimesheetOut)
def timesheet(
    period: str = "weekly",
    day: date | None = None,
    user_id: int | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    now = utcnow()
    target_day = day or _local_date_from_utc(now)
    target_user_id = _resolve_target_user_id(current_user, user_id)

    if period not in {"daily", "weekly", "monthly"}:
        raise HTTPException(status_code=400, detail="Invalid period")

    if period == "daily":
        start_date = target_day
        end_date = target_day
    elif period == "monthly":
        start_date = date(target_day.year, target_day.month, 1)
        end_date = date(target_day.year, target_day.month, calendar.monthrange(target_day.year, target_day.month)[1])
    else:
        start_date, end_date = _week_bounds(target_day)

    start_dt, end_dt = _local_period_bounds_utc(start_date, end_date)
    entries = _entries_overlapping_period(db, target_user_id, start_dt, end_dt)

    clock_hours = 0.0
    clocked_dates: set[date] = set()
    for entry in entries:
        net = _entry_metrics_for_period(db, entry, start_dt, end_dt, now=now)["net_hours"]
        clock_hours += net
        entry_local = _local_date_from_utc(entry.clock_in)
        clocked_dates.add(entry_local)

    target_user = db.get(User, target_user_id)
    required_daily = _sanitized_required_daily_hours(target_user) if target_user else 8.0
    absence_hours = _credited_absence_hours_for_period(
        db, target_user_id, start_date, end_date, required_daily, clocked_dates
    )

    return TimesheetOut(
        user_id=target_user_id,
        total_hours=round(clock_hours + absence_hours, 2),
        period_start=start_date,
        period_end=end_date,
    )


@router.get("/timesheet/export.csv")
def export_timesheet_csv(
    period: str = "weekly",
    day: date | None = None,
    user_id: int | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    now = utcnow()
    target_day = day or _local_date_from_utc(now)
    target_user_id = _resolve_target_user_id(current_user, user_id)

    if period not in {"daily", "weekly"}:
        raise HTTPException(status_code=400, detail="Invalid period")

    if period == "daily":
        start_date = target_day
        end_date = target_day
    else:
        start_date, end_date = _week_bounds(target_day)

    start_dt, end_dt = _local_period_bounds_utc(start_date, end_date)
    entries = _entries_overlapping_period(db, target_user_id, start_dt, end_dt)
    entries.sort(key=lambda entry: entry.clock_in, reverse=True)

    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "clock_entry_id",
            "user_id",
            "clock_in",
            "clock_out",
            "is_open",
            "break_hours",
            "required_break_hours",
            "deducted_break_hours",
            "net_hours",
        ]
    )
    for entry in entries:
        metrics = _entry_metrics(db, entry, now=now)
        writer.writerow(
            [
                entry.id,
                entry.user_id,
                entry.clock_in.isoformat(),
                entry.clock_out.isoformat() if entry.clock_out else "",
                "true" if entry.clock_out is None else "false",
                metrics["break_hours"],
                metrics["required_break_hours"],
                metrics["deducted_break_hours"],
                metrics["net_hours"],
            ]
        )

    filename = f"timesheet-{target_user_id}-{start_date.isoformat()}-{period}.csv"
    return Response(
        content=buffer.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/timesheet/export.xlsx")
def export_timesheet_xlsx(
    month: str | None = None,   # format: "YYYY-MM", defaults to current month
    user_id: int | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export a monthly timesheet as Excel, matching the Stundenliste template."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    now_local = _local_date_from_utc(utcnow())
    if month:
        try:
            year, mon = (int(x) for x in month.split("-"))
        except (ValueError, AttributeError):
            raise HTTPException(status_code=400, detail="month must be YYYY-MM")
    else:
        year, mon = now_local.year, now_local.month

    target_user_id = _resolve_target_user_id(current_user, user_id)
    target_user = db.get(User, target_user_id)
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    required_daily = _sanitized_required_daily_hours(target_user)
    month_start = date(year, mon, 1)
    month_end = date(year, mon, calendar.monthrange(year, mon)[1])

    # Fetch all clock entries for the month
    start_dt, end_dt = _local_period_bounds_utc(month_start, month_end)
    entries = _entries_overlapping_period(db, target_user_id, start_dt, end_dt)
    now_utc = utcnow()

    # Pre-compute per-entry metrics using MONTH bounds (consistent with timesheet endpoint).
    # Using month bounds ensures break deductions are applied per-entry, not per-day-slice,
    # so the totals match the overview numbers shown in the UI.
    entry_metrics_cache: dict[int, dict] = {
        entry.id: _entry_metrics_for_period(db, entry, start_dt, end_dt, now=now_utc)
        for entry in entries
    }

    # Build day-indexed lookup: date → list of clock entries.
    # Entries whose clock_in falls before the month start (cross-month overnight entries)
    # are attributed to the first day of the month so they're not silently dropped.
    day_entries: dict[date, list] = {}
    for entry in entries:
        entry_date = _local_date_from_utc(entry.clock_in)
        if entry_date < month_start:
            entry_date = month_start
        elif entry_date > month_end:
            continue
        day_entries.setdefault(entry_date, []).append(entry)

    # Fetch absences
    school_abs = db.scalars(
        select(SchoolAbsence).where(
            SchoolAbsence.user_id == target_user_id,
            SchoolAbsence.start_date <= month_end,
            SchoolAbsence.end_date >= month_start,
        )
    ).all()
    vacations = db.scalars(
        select(VacationRequest).where(
            VacationRequest.user_id == target_user_id,
            VacationRequest.status == "approved",
            VacationRequest.start_date <= month_end,
            VacationRequest.end_date >= month_start,
        )
    ).all()

    # Build absence lookup: date → (label, counts_as_hours)
    day_absence: dict[date, tuple[str, bool]] = {}
    for va in vacations:
        cur = max(va.start_date, month_start)
        while cur <= min(va.end_date, month_end):
            day_absence[cur] = ("Urlaub", True)
            cur += timedelta(days=1)
    for ab in school_abs:
        cur = max(ab.start_date, month_start)
        while cur <= min(ab.end_date, month_end):
            meta = _ABSENCE_TYPE_MAP.get(ab.absence_type, {})
            day_absence[cur] = (
                ab.title or meta.get("label_de", ab.absence_type),
                ab.counts_as_hours,
            )
            cur += timedelta(days=1)

    # Calculate totals
    total_worked = 0.0
    total_break = 0.0
    workdays_in_month = sum(1 for d in range(1, calendar.monthrange(year, mon)[1] + 1) if date(year, mon, d).weekday() < 5)
    total_required = round(workdays_in_month * required_daily, 2)

    # ── Build workbook ──────────────────────────────────────────────────────
    wb = Workbook()

    # Helper styles
    def bold_font(size=10):
        return Font(name="Arial", size=size, bold=True)

    def normal_font(size=10):
        return Font(name="Arial", size=size)

    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    month_name_de = [
        "Januar","Februar","März","April","Mai","Juni",
        "Juli","August","September","Oktober","November","Dezember"
    ][mon - 1]

    # ── Sheet 1: Übersichtsblatt ────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Übersichtsblatt"

    # Header row
    ws_sum["A1"] = f"Übersichtsblatt {month_name_de} {year}"
    ws_sum["A1"].font = bold_font(12)
    ws_sum["I1"] = f"Stand: {now_local.strftime('%d.%m.%Y')}"
    ws_sum["I1"].font = normal_font(9)
    ws_sum["I1"].alignment = right_align

    headers_sum = [
        "Mitarbeiter", "Personalnummer", "Soll-Stunden", "Ist-Stunden",
        "Differenz", "Urlaubstage:\nGenommen", "Urlaubstage:\nGeplant",
        "Urlaubstage:\nOffen", "Überstunden",
    ]
    for col, h in enumerate(headers_sum, 1):
        cell = ws_sum.cell(row=2, column=col, value=h)
        cell.font = bold_font()
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Count vacation days used this month
    vacation_days_month = sum(
        1 for d in range(1, calendar.monthrange(year, mon)[1] + 1)
        if date(year, mon, d) in day_absence and day_absence[date(year, mon, d)][0] == "Urlaub"
        and date(year, mon, d).weekday() < 5
    )

    # Calculate actual month hours (mirrors the detail-sheet row logic)
    overview_holidays = _nrw_public_holidays(year)
    month_is_hours = 0.0
    for d in range(1, calendar.monthrange(year, mon)[1] + 1):
        cur_date = date(year, mon, d)
        if cur_date.weekday() >= 5:
            continue
        if cur_date in overview_holidays:
            month_is_hours += required_daily
        elif cur_date in day_entries:
            for entry in day_entries[cur_date]:
                month_is_hours += entry_metrics_cache[entry.id]["net_hours"]
        elif cur_date in day_absence:
            _, cah = day_absence[cur_date]
            if cah:
                month_is_hours += required_daily

    month_is_hours = round(month_is_hours, 2)
    diff = round(month_is_hours - total_required, 2)

    name_parts = target_user.full_name.rsplit(" ", 1)
    display_name = f"{name_parts[-1]}, {name_parts[0]}" if len(name_parts) == 2 else target_user.full_name
    personal_nr = f"{target_user.id:05d}"

    row_data = [display_name, personal_nr, total_required, month_is_hours, diff,
                vacation_days_month, 0, -vacation_days_month, 0]
    for col, val in enumerate(row_data, 1):
        cell = ws_sum.cell(row=3, column=col, value=val)
        cell.font = normal_font()
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="right" if col > 2 else "left", vertical="center")

    # Summe row
    ws_sum.cell(row=4, column=1, value="Summe:").font = bold_font()
    for col, val in enumerate(row_data, 1):
        cell = ws_sum.cell(row=4, column=col + 0 if col == 1 else col, value=(None if col == 1 else val))
        if col > 1:
            cell.font = bold_font()
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right", vertical="center")

    col_widths = [22, 14, 12, 12, 12, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w
    ws_sum.row_dimensions[2].height = 28

    # ── Sheet 2: Employee detail ────────────────────────────────────────────
    sheet_name = f"{name_parts[-1]}_{name_parts[0]}" if len(name_parts) == 2 else target_user.full_name.replace(" ", "_")
    sheet_name = sheet_name[:31]  # Excel sheet name limit
    ws = wb.create_sheet(title=sheet_name)

    # Title
    ws["A1"] = f"Stundenliste mit Überstunden - {month_name_de} {year}"
    ws["A1"].font = bold_font(12)
    ws.merge_cells("A1:E1")

    # Employee info block
    ws["F1"] = "Name, Vorname"
    ws["F1"].font = bold_font()
    ws["H1"] = "Abteilung"
    ws["H1"].font = bold_font()
    ws["J1"] = "Position"
    ws["J1"].font = bold_font()

    ws["F2"] = display_name
    ws["F2"].font = normal_font()

    ws["F4"] = "Stand: " + now_local.strftime("%d.%m.%Y")
    ws["F4"].font = normal_font(9)
    ws["F5"] = "Personal-Nr."
    ws["F5"].font = bold_font()
    ws["H5"] = "Kostenstelle"
    ws["H5"].font = bold_font()
    ws["J5"] = "Vergütungsart"
    ws["J5"].font = bold_font()
    ws["F6"] = personal_nr
    ws["H6"] = "-"
    ws["J6"] = "Lohn"

    # Column headers (row 6 → row 7 because info block occupies rows 1-6)
    DATA_START_ROW = 7
    col_headers = ["Datum", "", "Startzeit", "Endzeit", "Pausenzeit",
                   "Soll-Stunden", "Ist-Stunden", "Differenz", "Anmerkungen", "Abwesenheit"]
    for col, h in enumerate(col_headers, 1):
        cell = ws.cell(row=DATA_START_ROW, column=col, value=h)
        cell.font = bold_font()
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Day rows
    row = DATA_START_ROW + 1
    total_worked_hours = 0.0
    total_break_hours = 0.0
    total_soll = 0.0  # accumulated from actual weekday Soll values

    month_holidays = _nrw_public_holidays(year)

    for d in range(1, calendar.monthrange(year, mon)[1] + 1):
        cur_date = date(year, mon, d)
        wd = cur_date.weekday()
        wd_abbr = _WEEKDAY_ABBR_DE[wd]
        date_str = cur_date.strftime("%d.%m.%Y")
        is_weekend = wd >= 5
        holiday_name = month_holidays.get(cur_date)

        if is_weekend:
            # Weekends: all zeros, no Soll
            row_vals = [wd_abbr, date_str, "-", "-", 0.0, 0.0, 0.0, 0.0, None, None]
        elif holiday_name:
            # Public holiday on a weekday — full hours credited automatically
            total_soll += required_daily
            total_worked_hours += required_daily
            row_vals = [wd_abbr, date_str, "-", "-", 0.0,
                        required_daily, required_daily, 0.0, holiday_name, "Feiertag"]
        elif cur_date in day_entries:
            day_net = 0.0
            day_break = 0.0
            first_in = None
            last_out = None
            for entry in sorted(day_entries[cur_date], key=lambda e: e.clock_in):
                metrics = entry_metrics_cache[entry.id]
                day_net += metrics["net_hours"]
                # deducted_break_hours = max(recorded, legally-required) — this is
                # the amount actually subtracted from gross time to produce net_hours.
                day_break += metrics["deducted_break_hours"]
                if first_in is None:
                    local_in = entry.clock_in.replace(tzinfo=timezone.utc).astimezone(_app_timezone())
                    first_in = local_in.strftime("%H:%M")
                if entry.clock_out:
                    local_out = entry.clock_out.replace(tzinfo=timezone.utc).astimezone(_app_timezone())
                    last_out = local_out.strftime("%H:%M")
            day_net = round(day_net, 2)
            day_break = round(day_break, 2)
            day_diff = round(day_net - required_daily, 2)
            total_soll += required_daily
            total_worked_hours += day_net
            total_break_hours += day_break
            row_vals = [wd_abbr, date_str, first_in or "-", last_out or "-",
                        day_break, required_daily, day_net, day_diff, None, None]
        elif cur_date in day_absence:
            absence_label, cah = day_absence[cur_date]
            credited = round(required_daily if cah else 0.0, 2)
            day_diff = round(credited - required_daily, 2)
            total_soll += required_daily
            total_worked_hours += credited
            row_vals = [wd_abbr, date_str, "-", "-", 0.0,
                        required_daily, credited, day_diff, absence_label, absence_label]
        else:
            # Empty weekday — no clock entry, no absence, no holiday
            day_diff = round(0.0 - required_daily, 2)
            total_soll += required_daily
            row_vals = [wd_abbr, date_str, "-", "-", 0.0,
                        required_daily, 0.0, day_diff, None, None]

        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = normal_font()
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right" if col > 2 else "left", vertical="center")

        # Grey-out weekends
        if is_weekend:
            grey_fill = PatternFill("solid", fgColor="F2F2F2")
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = grey_fill
        # Tint public holiday rows
        elif holiday_name:
            holiday_fill = PatternFill("solid", fgColor="FEF3C7")
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = holiday_fill

        row += 1

    # Totals row
    total_worked_hours = round(total_worked_hours, 2)
    total_break_hours = round(total_break_hours, 2)
    total_soll = round(total_soll, 2)
    month_diff = round(total_worked_hours - total_soll, 2)
    total_label = f"Summen vom {month_start.strftime('%d.%m.%Y')} bis {month_end.strftime('%d.%m.%Y')}"
    total_row_vals = [total_label, None, None, None, total_break_hours,
                      total_soll, total_worked_hours, month_diff, None, 0]
    for col, val in enumerate(total_row_vals, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = bold_font()
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="right" if col > 1 else "left", vertical="center")

    # Footer summary block
    row += 2
    ws.cell(row=row, column=1, value="Aufschlüsselung der Abwesenheiten:").font = bold_font()
    row += 2

    ws.cell(row=row, column=1, value="Urlaubsübertrag aus dem Vorjahr:").font = normal_font()
    ws.cell(row=row, column=4, value=0).font = normal_font()
    ws.cell(row=row, column=6, value="Überstunden diesen Monat:").font = normal_font()
    ws.cell(row=row, column=8, value=month_diff).font = normal_font()
    row += 1

    ws.cell(row=row, column=1, value="Jahresurlaub (inkl. Übertrag):").font = normal_font()
    ws.cell(row=row, column=4, value=0).font = normal_font()
    ws.cell(row=row, column=6, value="Überstundensaldo Monatsbeginn:").font = normal_font()
    ws.cell(row=row, column=8, value=0).font = normal_font()
    row += 1

    ws.cell(row=row, column=1, value="Genutzte Urlaubstage diesen Monat:").font = normal_font()
    ws.cell(row=row, column=4, value=vacation_days_month).font = normal_font()
    ws.cell(row=row, column=6, value="Korrigierter Überstundensaldo:").font = normal_font()
    ws.cell(row=row, column=8, value=0).font = normal_font()
    row += 1

    ws.cell(row=row, column=1, value="Genutzte Urlaubstage dieses Jahr:").font = normal_font()
    ws.cell(row=row, column=4, value=vacation_days_month).font = normal_font()
    ws.cell(row=row, column=6, value="Ausbezahlte Stunden:").font = normal_font()
    ws.cell(row=row, column=8, value=0).font = normal_font()
    row += 1

    ws.cell(row=row, column=1, value="Verbleibende Urlaubstage dieses Jahr:").font = normal_font()
    ws.cell(row=row, column=4, value=-vacation_days_month).font = normal_font()
    ws.cell(row=row, column=6, value="Überstundensaldo Monatsende:").font = normal_font()
    ws.cell(row=row, column=8, value=month_diff).font = normal_font()

    # Column widths for detail sheet
    detail_widths = [10, 12, 10, 10, 10, 12, 12, 12, 22, 12]
    for i, w in enumerate(detail_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Serialize
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = display_name.replace(", ", "_").replace(" ", "_")
    filename = f"Stundenliste_{safe_name}_{year}_{mon:02d}.xlsx"
    headers_resp = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )


@router.get("/entries", response_model=list[TimeEntryOut])
def list_entries(
    period: str = "weekly",
    day: date | None = None,
    user_id: int | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    target_day = day or _local_date_from_utc(utcnow())
    target_user_id = _resolve_target_user_id(current_user, user_id)

    if period not in {"daily", "weekly"}:
        raise HTTPException(status_code=400, detail="Invalid period")

    if period == "daily":
        start_date = target_day
        end_date = target_day
    else:
        start_date, end_date = _week_bounds(target_day)

    start_dt, end_dt = _local_period_bounds_utc(start_date, end_date)
    entries = _entries_overlapping_period(db, target_user_id, start_dt, end_dt)
    entries.sort(key=lambda entry: entry.clock_in, reverse=True)
    now = utcnow()
    return [_entry_out(db, entry, now=now) for entry in entries]


@router.patch("/entries/{clock_entry_id}", response_model=TimeEntryOut)
def update_entry(
    clock_entry_id: int,
    payload: TimeEntryUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    entry = db.get(ClockEntry, clock_entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Clock entry not found")
    if entry.user_id != current_user.id and not _is_time_manager(current_user):
        raise HTTPException(status_code=403, detail="Not allowed")

    if payload.clock_out and payload.clock_out < payload.clock_in:
        raise HTTPException(status_code=400, detail="clock_out must be after clock_in")

    worked_minutes = _hours_between(payload.clock_in, payload.clock_out) * 60 if payload.clock_out else None
    if worked_minutes is not None and payload.break_minutes > worked_minutes:
        raise HTTPException(status_code=400, detail="break_minutes exceeds worked duration")

    entry.clock_in = payload.clock_in
    entry.clock_out = payload.clock_out
    db.add(entry)
    db.flush()

    existing_breaks = db.scalars(select(BreakEntry).where(BreakEntry.clock_entry_id == entry.id)).all()
    for existing in existing_breaks:
        db.delete(existing)

    if payload.break_minutes > 0:
        break_end = payload.clock_out or (payload.clock_in + timedelta(minutes=payload.break_minutes))
        break_start = break_end - timedelta(minutes=payload.break_minutes)
        if break_start < payload.clock_in:
            break_start = payload.clock_in
        db.add(BreakEntry(clock_entry_id=entry.id, break_start=break_start, break_end=break_end))

    db.commit()
    db.refresh(entry)
    return _entry_out(db, entry)
