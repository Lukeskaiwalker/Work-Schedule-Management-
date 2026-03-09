from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
import csv
import re
import unicodedata
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.entities import Project, ProjectFinance


PROJECT_NUMBER_KEYS = {
    "project_number",
    "projektnummer",
    "projekt_nummer",
    "projekt_nr",
    "project_id",
    "project_no",
    "project_number_id",
    "auftragsnummer",
    "vorgangsnummer",
    "nr",
    "nummer",
    "lfd_nr",
}

PROJECT_NAME_KEYS = {
    "name",
    "project_name",
    "projektname",
    "projekt",
    "bezeichnung",
    "titel",
    "projekt_anfrage",
}

DESCRIPTION_KEYS = {
    "description",
    "beschreibung",
    "bemerkung",
    "kommentar",
}

LAST_STATE_KEYS = {
    "last_state",
    "notiz",
    "notizen",
    "notes",
    "notizen_beschreibung",
}

LAST_STATUS_AT_KEYS = {
    "last_status_at",
    "letzter_status_datum",
    "last_status_date",
    "status_updated_at",
}

STATUS_KEYS = {"status", "projektstatus", "project_status", "aktueller_status"}

CUSTOMER_NAME_KEYS = {
    "customer_name",
    "kunde",
    "kundenname",
    "client",
    "auftraggeber",
    "firma",
}

CUSTOMER_ADDRESS_KEYS = {
    "customer_address",
    "adresse",
    "anschrift",
    "kundenadresse",
}

CONSTRUCTION_SITE_ADDRESS_KEYS = {
    "construction_site_address",
    "baustellenadresse",
    "site_address",
    "baustelle",
}

CUSTOMER_CONTACT_KEYS = {
    "customer_contact",
    "ansprechpartner",
    "kontaktperson",
    "kontakt",
}

CUSTOMER_EMAIL_KEYS = {"customer_email", "email", "e_mail", "mail"}

CUSTOMER_PHONE_KEYS = {"customer_phone", "telefon", "telefonnummer", "phone", "tel", "mobil"}

FINANCE_ORDER_VALUE_NET_KEYS = {
    "order_value_net",
    "auftragswert_netto",
    "auftragswert_net",
    "auftragswert",
}
FINANCE_DOWN_PAYMENT_35_KEYS = {
    "down_payment_35",
    "anzahlung_35",
    "35_anzahlung",
    "anzahlung35",
}
FINANCE_MAIN_COMPONENTS_50_KEYS = {
    "main_components_50",
    "hauptkomponenten_50",
    "50_hauptkomponenten",
    "hauptkomponenten50",
}
FINANCE_FINAL_INVOICE_15_KEYS = {
    "final_invoice_15",
    "schlussrechnung_15",
    "15_schlussrechnung",
    "schlussrechnung15",
}
FINANCE_PLANNED_COSTS_KEYS = {"planned_costs", "geplante_kosten"}
FINANCE_ACTUAL_COSTS_KEYS = {"actual_costs", "tatsachliche_kosten", "tatsaechliche_kosten"}
FINANCE_CONTRIBUTION_MARGIN_KEYS = {"contribution_margin", "deckungsbeitrag"}
FINANCE_PLANNED_HOURS_TOTAL_KEYS = {"planned_hours_total", "geplante_stunden", "projektstunden_geplant"}

FINANCE_FIELD_KEY_MAP: tuple[tuple[str, set[str]], ...] = (
    ("order_value_net", FINANCE_ORDER_VALUE_NET_KEYS),
    ("down_payment_35", FINANCE_DOWN_PAYMENT_35_KEYS),
    ("main_components_50", FINANCE_MAIN_COMPONENTS_50_KEYS),
    ("final_invoice_15", FINANCE_FINAL_INVOICE_15_KEYS),
    ("planned_costs", FINANCE_PLANNED_COSTS_KEYS),
    ("actual_costs", FINANCE_ACTUAL_COSTS_KEYS),
    ("contribution_margin", FINANCE_CONTRIBUTION_MARGIN_KEYS),
    ("planned_hours_total", FINANCE_PLANNED_HOURS_TOTAL_KEYS),
)


@dataclass
class RowData:
    original: dict[str, Any]
    normalized: dict[str, Any]
    sheet_name: str
    row_number: int
    sort_index: int


@dataclass
class ImportStats:
    created: int = 0
    updated: int = 0
    temporary_numbers: int = 0
    processed_rows: int = 0
    duplicates_skipped: int = 0
    skipped_project_fields: int = 0
    skipped_finance_fields: int = 0


def _normalize_key(value: str, fallback_index: int | None = None) -> str:
    text = str(value or "").strip()
    if not text and fallback_index is not None:
        text = f"column_{fallback_index + 1}"
    ascii_text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^a-zA-Z0-9]+", "_", ascii_text).strip("_").lower()
    if not normalized and fallback_index is not None:
        return f"column_{fallback_index + 1}"
    return normalized


def _json_safe(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        trimmed = value.strip()
        return trimmed if trimmed else None
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _first_value(row: dict[str, Any], candidates: set[str]) -> Any:
    for key in candidates:
        value = row.get(key)
        if not _is_empty(value):
            return value
    return None


def _as_clean_string(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _project_status_value(value: Any) -> str:
    if _is_empty(value):
        return "active"
    return re.sub(r"\s+", " ", _as_clean_string(value))


def _parse_float(value: Any) -> float | None:
    if _is_empty(value):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    raw = _as_clean_string(value)
    if not raw:
        return None
    normalized = (
        raw.replace("\u00a0", " ")
        .replace(" ", "")
        .replace("EUR", "")
        .replace("eur", "")
        .replace("€", "")
    )
    if not normalized:
        return None

    if "," in normalized and "." in normalized:
        if normalized.rfind(",") > normalized.rfind("."):
            normalized = normalized.replace(".", "").replace(",", ".")
        else:
            normalized = normalized.replace(",", "")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")

    try:
        return float(normalized)
    except ValueError:
        return None


def _parse_datetime(value: Any) -> datetime | None:
    if _is_empty(value):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            parsed = from_excel(value)
            if isinstance(parsed, datetime):
                return parsed.replace(tzinfo=None)
            if isinstance(parsed, date):
                return datetime.combine(parsed, time.min)
        except Exception:
            return None
        return None

    text = _as_clean_string(value)
    if not text:
        return None
    iso_candidate = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(iso_candidate)
        if parsed.tzinfo is not None:
            parsed = parsed.astimezone().replace(tzinfo=None)
        return parsed
    except ValueError:
        pass

    formats = (
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _detect_header_row(rows: list[tuple[Any, ...]]) -> int:
    best_row_index = 1
    best_non_empty = 0
    for idx, row in enumerate(rows, start=1):
        non_empty = sum(0 if _is_empty(cell) else 1 for cell in row)
        if non_empty > best_non_empty:
            best_non_empty = non_empty
            best_row_index = idx
    if best_non_empty < 2:
        raise ValueError("Could not detect header row in Excel file")
    return best_row_index


def _is_repeated_header_row(values: list[Any], normalized_headers: list[str]) -> bool:
    limited = values[: len(normalized_headers)]
    non_empty_indexes = [idx for idx, value in enumerate(limited) if not _is_empty(value)]
    if len(non_empty_indexes) < 2:
        return False

    matches = 0
    for idx in non_empty_indexes:
        cell_key = _normalize_key(_as_clean_string(limited[idx]), idx)
        if cell_key == normalized_headers[idx]:
            matches += 1
    return matches >= max(2, int(len(non_empty_indexes) * 0.7))


def _sheet_has_project_identity_columns(normalized_headers: list[str]) -> bool:
    header_set = set(normalized_headers)
    identity_keys = PROJECT_NUMBER_KEYS | PROJECT_NAME_KEYS | CUSTOMER_NAME_KEYS
    return bool(header_set & identity_keys)


def _build_headers(header_tuple: tuple[Any, ...], max_columns: int) -> tuple[list[str], list[str]]:
    original_headers: list[str] = []
    normalized_headers: list[str] = []
    seen_original: dict[str, int] = {}
    seen_normalized: dict[str, int] = {}

    for idx, raw_header in enumerate(header_tuple[:max_columns]):
        base_original = str(raw_header).strip() if raw_header is not None else f"Column {idx + 1}"
        if not base_original:
            base_original = f"Column {idx + 1}"
        original_count = seen_original.get(base_original, 0) + 1
        seen_original[base_original] = original_count
        original_header = base_original if original_count == 1 else f"{base_original} ({original_count})"

        base_normalized = _normalize_key(base_original, idx)
        normalized_count = seen_normalized.get(base_normalized, 0) + 1
        seen_normalized[base_normalized] = normalized_count
        normalized_header = base_normalized if normalized_count == 1 else f"{base_normalized}_{normalized_count}"

        original_headers.append(original_header)
        normalized_headers.append(normalized_header)
    return original_headers, normalized_headers


def _max_used_column_index(sheet: Any, *, min_row: int) -> int:
    max_idx = -1
    for values in sheet.iter_rows(min_row=min_row, values_only=True):
        row_values = list(values)
        for idx in range(len(row_values) - 1, -1, -1):
            if not _is_empty(row_values[idx]):
                if idx > max_idx:
                    max_idx = idx
                break
    return max_idx


def _load_sheet_rows(sheet: Any, *, sheet_name: str, sheet_order: int) -> list[RowData]:
    scan_rows = list(sheet.iter_rows(min_row=1, max_row=50, values_only=True))
    if not scan_rows:
        return []

    header_row_index = _detect_header_row(scan_rows)
    header_tuple = next(sheet.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True))
    max_used_idx = _max_used_column_index(sheet, min_row=header_row_index)
    if max_used_idx < 0:
        return []

    original_headers, normalized_headers = _build_headers(header_tuple, max_used_idx + 1)
    if not _sheet_has_project_identity_columns(normalized_headers):
        return []

    rows: list[RowData] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=header_row_index + 1, values_only=True), start=header_row_index + 1):
        source_values = list(values)
        if len(source_values) < len(original_headers):
            source_values.extend([None] * (len(original_headers) - len(source_values)))
        source_values = source_values[: len(original_headers)]

        if _is_repeated_header_row(source_values, normalized_headers):
            continue

        original_map: dict[str, Any] = {}
        normalized_map: dict[str, Any] = {}
        empty_row = True
        for idx, original_header in enumerate(original_headers):
            value = _json_safe(source_values[idx] if idx < len(source_values) else None)
            original_map[original_header] = value
            normalized_map[normalized_headers[idx]] = value
            if not _is_empty(value):
                empty_row = False

        if not empty_row:
            rows.append(
                RowData(
                    original=original_map,
                    normalized=normalized_map,
                    sheet_name=sheet_name,
                    row_number=row_number,
                    sort_index=(sheet_order * 1_000_000) + row_number,
                )
            )
    return rows


def _load_rows(file_path: Path, sheet_name: str | None = None) -> list[RowData]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel imports") from exc

    workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        sheet_names = [sheet_name]
    else:
        sheet_names = list(workbook.sheetnames)

    rows: list[RowData] = []
    for sheet_order, name in enumerate(sheet_names):
        rows.extend(_load_sheet_rows(workbook[name], sheet_name=name, sheet_order=sheet_order))
    return rows


def _load_csv_rows(file_path: Path) -> list[RowData]:
    rows: list[RowData] = []
    with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        all_rows = list(reader)
    if not all_rows:
        return rows
    header = all_rows[0]
    original_headers, normalized_headers = _build_headers(tuple(header), len(header))
    if not _sheet_has_project_identity_columns(normalized_headers):
        return rows

    for row_number, values in enumerate(all_rows[1:], start=2):
        source_values = list(values)
        if len(source_values) < len(original_headers):
            source_values.extend([""] * (len(original_headers) - len(source_values)))
        source_values = source_values[: len(original_headers)]

        original_map: dict[str, Any] = {}
        normalized_map: dict[str, Any] = {}
        empty_row = True
        for idx, original_header in enumerate(original_headers):
            value = _json_safe(source_values[idx] if idx < len(source_values) else None)
            original_map[original_header] = value
            normalized_map[normalized_headers[idx]] = value
            if not _is_empty(value):
                empty_row = False
        if empty_row:
            continue
        rows.append(
            RowData(
                original=original_map,
                normalized=normalized_map,
                sheet_name="csv",
                row_number=row_number,
                sort_index=row_number,
            )
        )
    return rows


def _build_fallback_identity(
    customer_value: Any,
    name_value: Any,
    customer_address_value: Any,
    construction_site_address_value: Any,
) -> str | None:
    identity_parts = [
        _normalize_key(_as_clean_string(customer_value)) if not _is_empty(customer_value) else "",
        _normalize_key(_as_clean_string(name_value)) if not _is_empty(name_value) else "",
        _normalize_key(_as_clean_string(customer_address_value)) if not _is_empty(customer_address_value) else "",
        _normalize_key(_as_clean_string(construction_site_address_value))
        if not _is_empty(construction_site_address_value)
        else "",
    ]
    if any(identity_parts):
        return f"fallback:{'|'.join(identity_parts)}"
    return None


def _row_fallback_identity(row: RowData) -> str | None:
    name_value = _first_value(row.normalized, PROJECT_NAME_KEYS)
    if _is_empty(name_value):
        name_value = _first_value(row.normalized, {"projekt_anfrage", "customer_name", "kunde"})
    customer_value = _first_value(row.normalized, CUSTOMER_NAME_KEYS)
    customer_address_value = _first_value(row.normalized, CUSTOMER_ADDRESS_KEYS)
    construction_site_address_value = _first_value(row.normalized, CONSTRUCTION_SITE_ADDRESS_KEYS)
    return _build_fallback_identity(
        customer_value,
        name_value,
        customer_address_value,
        construction_site_address_value,
    )


def _project_fallback_identity(project: Project) -> str | None:
    return _build_fallback_identity(
        project.customer_name,
        project.name,
        project.customer_address,
        project.construction_site_address,
    )


def _row_identity_key(row: RowData) -> str:
    project_number = _first_value(row.normalized, PROJECT_NUMBER_KEYS)
    if not _is_empty(project_number):
        return f"number:{_normalize_key(_as_clean_string(project_number))}"

    fallback_identity = _row_fallback_identity(row)
    if fallback_identity:
        return fallback_identity
    return f"row:{row.sheet_name}:{row.row_number}"


def _row_has_project_number(row: RowData) -> bool:
    return not _is_empty(_first_value(row.normalized, PROJECT_NUMBER_KEYS))


def _row_status_datetime(row: RowData) -> datetime | None:
    return _parse_datetime(_first_value(row.normalized, LAST_STATUS_AT_KEYS))


def _pick_preferred_row(current: RowData, candidate: RowData) -> RowData:
    current_dt = _row_status_datetime(current)
    candidate_dt = _row_status_datetime(candidate)
    if candidate_dt and current_dt:
        if candidate_dt > current_dt:
            return candidate
        if candidate_dt < current_dt:
            return current
    elif candidate_dt and not current_dt:
        return candidate
    elif current_dt and not candidate_dt:
        return current

    if _row_has_project_number(candidate) and not _row_has_project_number(current):
        return candidate
    if _row_has_project_number(current) and not _row_has_project_number(candidate):
        return current
    return candidate if candidate.sort_index >= current.sort_index else current


def _deduplicate_rows(rows: list[RowData]) -> tuple[list[RowData], int]:
    by_identity: dict[str, RowData] = {}
    duplicates = 0
    for row in rows:
        key = _row_identity_key(row)
        existing = by_identity.get(key)
        if existing is None:
            by_identity[key] = row
            continue
        by_identity[key] = _pick_preferred_row(existing, row)
        duplicates += 1
    deduped = sorted(by_identity.values(), key=lambda item: item.sort_index)
    return deduped, duplicates


def _next_temporary_number(reserved_numbers: set[str], counter: int) -> tuple[str, int]:
    current = counter
    while True:
        candidate = f"T{current:05d}"
        if candidate not in reserved_numbers:
            reserved_numbers.add(candidate)
            return candidate, current + 1
        current += 1


def _string_or_none(value: Any) -> str | None:
    if _is_empty(value):
        return None
    return _as_clean_string(value)


def _set_if_missing(instance: object, field_name: str, value: Any) -> tuple[bool, bool]:
    if value is None:
        return False, False
    current = getattr(instance, field_name)
    if not _is_empty(current):
        return False, True
    setattr(instance, field_name, value)
    return True, False


def _finance_values_from_row(row_normalized: dict[str, Any]) -> dict[str, float | None]:
    values: dict[str, float | None] = {}
    for field_name, keys in FINANCE_FIELD_KEY_MAP:
        values[field_name] = _parse_float(_first_value(row_normalized, keys))
    return values


def _merge_extra_attributes(existing: dict[str, Any], incoming: dict[str, Any]) -> tuple[dict[str, Any], bool, int]:
    merged = dict(existing or {})
    changed = False
    skipped = 0
    for key, value in incoming.items():
        if _is_empty(value):
            continue
        if key not in merged or _is_empty(merged.get(key)):
            merged[key] = value
            changed = True
        else:
            skipped += 1
    return merged, changed, skipped


def _import_projects_from_rows(
    db: Session,
    rows: list[RowData],
    *,
    source_label: str | None = None,
) -> ImportStats:
    stats = ImportStats(processed_rows=len(rows))

    if not rows:
        return stats

    rows, duplicates_skipped = _deduplicate_rows(rows)
    stats.duplicates_skipped = duplicates_skipped

    existing_projects = db.scalars(select(Project)).all()
    existing_finances = db.scalars(select(ProjectFinance)).all()
    projects_by_number = {project.project_number: project for project in existing_projects}
    finances_by_project_id = {finance.project_id: finance for finance in existing_finances}
    projects_by_fallback: dict[str, Project] = {}
    for project in existing_projects:
        fallback_identity = _project_fallback_identity(project)
        if fallback_identity:
            projects_by_fallback[fallback_identity] = project
    reserved_numbers = set(projects_by_number.keys())
    temporary_counter = 1

    for row in rows:
        provided_project_number = _first_value(row.normalized, PROJECT_NUMBER_KEYS)
        project_number: str
        fallback_identity = _row_fallback_identity(row)
        project: Project | None = None
        project_number_changed = False

        if _is_empty(provided_project_number):
            if fallback_identity and fallback_identity in projects_by_fallback:
                project = projects_by_fallback[fallback_identity]
                project_number = project.project_number
                reserved_numbers.add(project_number)
            elif not fallback_identity:
                continue
            else:
                project_number, temporary_counter = _next_temporary_number(reserved_numbers, temporary_counter)
                stats.temporary_numbers += 1
        else:
            project_number = _as_clean_string(provided_project_number)
            reserved_numbers.add(project_number)
            project = projects_by_number.get(project_number)
            if project is None and fallback_identity and fallback_identity in projects_by_fallback:
                fallback_project = projects_by_fallback[fallback_identity]
                project = fallback_project
                if fallback_project.project_number != project_number and project_number not in projects_by_number:
                    old_number = fallback_project.project_number
                    fallback_project.project_number = project_number
                    project_number_changed = True
                    projects_by_number.pop(old_number, None)
                    reserved_numbers.discard(old_number)

        name_value = _first_value(row.normalized, PROJECT_NAME_KEYS)
        if _is_empty(name_value):
            name_value = _first_value(row.normalized, {"projekt_anfrage", "customer_name", "kunde"})
        name_from_import = _as_clean_string(name_value) if not _is_empty(name_value) else None
        name = name_from_import or f"Project {project_number}"

        status_value = _first_value(row.normalized, STATUS_KEYS)
        status = _project_status_value(status_value)

        description_value = _first_value(row.normalized, DESCRIPTION_KEYS)
        last_state_value = _first_value(row.normalized, LAST_STATE_KEYS)
        last_status_at_value = _first_value(row.normalized, LAST_STATUS_AT_KEYS)
        last_status_at = _parse_datetime(last_status_at_value)
        customer_name = _first_value(row.normalized, CUSTOMER_NAME_KEYS)
        customer_address = _first_value(row.normalized, CUSTOMER_ADDRESS_KEYS)
        construction_site_address = _first_value(row.normalized, CONSTRUCTION_SITE_ADDRESS_KEYS)
        customer_contact = _first_value(row.normalized, CUSTOMER_CONTACT_KEYS)
        customer_email = _first_value(row.normalized, CUSTOMER_EMAIL_KEYS)
        customer_phone = _first_value(row.normalized, CUSTOMER_PHONE_KEYS)
        finance_values = _finance_values_from_row(row.normalized)

        extra_attributes = {key: value for key, value in row.original.items() if not _is_empty(value)}
        if source_label:
            extra_attributes["_import_source"] = source_label
        extra_attributes["_import_sheet"] = row.sheet_name
        extra_attributes["_import_row"] = row.row_number

        if project is None:
            project = projects_by_number.get(project_number)
        if project is None:
            project = Project(
                project_number=project_number,
                name=name,
                description=_string_or_none(description_value),
                status=status,
                last_state=_string_or_none(last_state_value),
                last_status_at=last_status_at,
                customer_name=_string_or_none(customer_name),
                customer_address=_string_or_none(customer_address),
                construction_site_address=_string_or_none(construction_site_address),
                customer_contact=_string_or_none(customer_contact),
                customer_email=_string_or_none(customer_email),
                customer_phone=_string_or_none(customer_phone),
                extra_attributes=extra_attributes,
            )
            db.add(project)
            if any(value is not None for value in finance_values.values()):
                if project.id is None:
                    db.flush()
                finance_row = ProjectFinance(
                    project_id=project.id,
                    order_value_net=finance_values["order_value_net"],
                    down_payment_35=finance_values["down_payment_35"],
                    main_components_50=finance_values["main_components_50"],
                    final_invoice_15=finance_values["final_invoice_15"],
                    planned_costs=finance_values["planned_costs"],
                    actual_costs=finance_values["actual_costs"],
                    contribution_margin=finance_values["contribution_margin"],
                    planned_hours_total=finance_values["planned_hours_total"],
                )
                db.add(finance_row)
                finances_by_project_id[project.id] = finance_row
            projects_by_number[project_number] = project
            if fallback_identity:
                projects_by_fallback[fallback_identity] = project
            stats.created += 1
            continue

        changed = project_number_changed
        applied, skipped = _set_if_missing(project, "name", name_from_import)
        changed = applied or changed
        stats.skipped_project_fields += int(skipped)
        if not _is_empty(status_value):
            applied, skipped = _set_if_missing(project, "status", status)
            changed = applied or changed
            stats.skipped_project_fields += int(skipped)
        applied, skipped = _set_if_missing(project, "description", _string_or_none(description_value))
        changed = applied or changed
        stats.skipped_project_fields += int(skipped)
        applied, skipped = _set_if_missing(project, "last_state", _string_or_none(last_state_value))
        changed = applied or changed
        stats.skipped_project_fields += int(skipped)
        applied, skipped = _set_if_missing(project, "last_status_at", last_status_at)
        changed = applied or changed
        stats.skipped_project_fields += int(skipped)
        applied, skipped = _set_if_missing(project, "customer_name", _string_or_none(customer_name))
        changed = applied or changed
        stats.skipped_project_fields += int(skipped)
        applied, skipped = _set_if_missing(project, "customer_address", _string_or_none(customer_address))
        changed = applied or changed
        stats.skipped_project_fields += int(skipped)
        applied, skipped = _set_if_missing(
            project,
            "construction_site_address",
            _string_or_none(construction_site_address),
        )
        changed = applied or changed
        stats.skipped_project_fields += int(skipped)
        applied, skipped = _set_if_missing(project, "customer_contact", _string_or_none(customer_contact))
        changed = applied or changed
        stats.skipped_project_fields += int(skipped)
        applied, skipped = _set_if_missing(project, "customer_email", _string_or_none(customer_email))
        changed = applied or changed
        stats.skipped_project_fields += int(skipped)
        applied, skipped = _set_if_missing(project, "customer_phone", _string_or_none(customer_phone))
        changed = applied or changed
        stats.skipped_project_fields += int(skipped)
        merged_extra, extra_changed, extra_skipped = _merge_extra_attributes(project.extra_attributes or {}, extra_attributes)
        if extra_changed:
            project.extra_attributes = merged_extra
            changed = True
        stats.skipped_project_fields += extra_skipped

        finance_changed = False
        if any(value is not None for value in finance_values.values()):
            if project.id is None:
                db.flush()
            finance_row = finances_by_project_id.get(project.id)
            if finance_row is None:
                finance_row = ProjectFinance(project_id=project.id)
                db.add(finance_row)
                finances_by_project_id[project.id] = finance_row
            for field_name, value in finance_values.items():
                if value is None:
                    continue
                applied, skipped = _set_if_missing(finance_row, field_name, value)
                if applied:
                    finance_changed = True
                stats.skipped_finance_fields += int(skipped)
            if finance_changed:
                db.add(finance_row)
        projects_by_number[project_number] = project
        if fallback_identity:
            projects_by_fallback[fallback_identity] = project
        if changed:
            db.add(project)
        if changed or finance_changed:
            stats.updated += 1

    db.commit()
    return stats


def import_projects_from_excel(
    db: Session,
    file_path: str,
    *,
    sheet_name: str | None = None,
    source_label: str | None = None,
) -> ImportStats:
    rows = _load_rows(Path(file_path), sheet_name=sheet_name)
    return _import_projects_from_rows(db, rows, source_label=source_label)


def import_projects_from_csv(
    db: Session,
    file_path: str,
    *,
    source_label: str | None = None,
) -> ImportStats:
    rows = _load_csv_rows(Path(file_path))
    return _import_projects_from_rows(db, rows, source_label=source_label)
