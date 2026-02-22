from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import func, select

from app.core.db import SessionLocal
from app.models.entities import Project
from app.services.project_import import import_projects_from_csv, import_projects_from_excel


def _write_workbook(path: Path, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)


def _write_csv(path: Path, rows: list[list[object]]) -> None:
    serialized = []
    for row in rows:
        serialized.append(",".join("" if value is None else str(value) for value in row))
    path.write_text("\n".join(serialized) + "\n", encoding="utf-8")


def test_import_projects_from_excel_preserves_columns_and_generates_temp_numbers(tmp_path: Path):
    file_path = tmp_path / "projects.xlsx"
    _write_workbook(
        file_path,
        [
            ["Projektname", "Projektnummer", "Kunde", "Adresse", "Neue Spalte"],
            ["PV Dach", "2026-9001", "Muster GmbH", "Berlin", "A"],
            ["Service Einsatz", None, "Klein AG", "Hamburg", "B"],
        ],
    )

    db = SessionLocal()
    try:
        stats = import_projects_from_excel(db, str(file_path), source_label=file_path.name)

        assert stats.processed_rows == 2
        assert stats.created == 2
        assert stats.updated == 0
        assert stats.temporary_numbers == 1

        projects = db.scalars(select(Project).order_by(Project.project_number.asc())).all()
        assert len(projects) == 2

        numbered = next(project for project in projects if project.project_number == "2026-9001")
        assert numbered.name == "PV Dach"
        assert numbered.customer_name == "Muster GmbH"
        assert numbered.extra_attributes["Neue Spalte"] == "A"
        assert numbered.extra_attributes["_import_source"] == "projects.xlsx"

        temporary = next(project for project in projects if project.project_number.startswith("T"))
        assert temporary.name == "Service Einsatz"
        assert temporary.customer_address == "Hamburg"
        assert temporary.extra_attributes["Neue Spalte"] == "B"
    finally:
        db.close()


def test_import_projects_from_excel_updates_existing_project(tmp_path: Path):
    file_path = tmp_path / "projects_update.xlsx"
    _write_workbook(
        file_path,
        [
            ["Projektname", "Projektnummer", "Status", "Notizen"],
            ["Alt", "2026-7777", "active", "first"],
        ],
    )

    db = SessionLocal()
    try:
        first = import_projects_from_excel(db, str(file_path), source_label=file_path.name)
        assert first.created == 1

        _write_workbook(
            file_path,
            [
                ["Projektname", "Projektnummer", "Status", "Notizen"],
                ["Neu", "2026-7777", "in_progress", "second"],
            ],
        )

        second = import_projects_from_excel(db, str(file_path), source_label=file_path.name)
        assert second.created == 0
        assert second.updated == 1

        project = db.scalars(select(Project).where(Project.project_number == "2026-7777")).first()
        assert project is not None
        assert project.name == "Neu"
        assert project.status == "in_progress"
        assert project.last_state == "second"
    finally:
        db.close()


def test_import_projects_from_excel_maps_nr_column_to_project_number(tmp_path: Path):
    file_path = tmp_path / "projects_nr.xlsx"
    _write_workbook(
        file_path,
        [
            ["Nr.", "Projekt/Anfrage", "Kunde"],
            [198, "PV Anlage", "SMPL Kunde"],
        ],
    )

    db = SessionLocal()
    try:
        stats = import_projects_from_excel(db, str(file_path), source_label=file_path.name)
        assert stats.created == 1
        assert stats.temporary_numbers == 0

        project = db.scalars(select(Project).where(Project.project_number == "198")).first()
        assert project is not None
        assert project.name == "PV Anlage"
        assert project.customer_name == "SMPL Kunde"
    finally:
        db.close()


def test_import_projects_from_excel_maps_german_status_and_notiz(tmp_path: Path):
    file_path = tmp_path / "projects_status_notiz.xlsx"
    _write_workbook(
        file_path,
        [
            ["Nr.", "Kunde", "Projektname", "Aktueller Status", "Notiz"],
            [321, "SMPL Kunde", "Solar Carport", "Kundentermin vereinbart", "Termin am Freitag bestaetigt"],
        ],
    )

    db = SessionLocal()
    try:
        stats = import_projects_from_excel(db, str(file_path), source_label=file_path.name)
        assert stats.created == 1

        project = db.scalars(select(Project).where(Project.project_number == "321")).first()
        assert project is not None
        assert project.status == "Kundentermin vereinbart"
        assert project.last_state == "Termin am Freitag bestaetigt"
        assert project.description is None
    finally:
        db.close()


def test_import_projects_from_excel_imports_last_status_datetime_and_deduplicates(tmp_path: Path):
    file_path = tmp_path / "projects_multi_sheet.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Angebote"
    ws1.append(["Nr.", "Kunde", "Projektname", "Aktueller Status", "Notiz", "Letzter Status Datum"])
    ws1.append([1001, "Muster GmbH", "PV Dach", "Angebot erstellen", "Erstkontakt", "20.02.2026 08:00"])
    ws1.append([None, "Ohne Nummer GmbH", "Serviceeinsatz", "Anfrage erhalten", "Ruft zurueck", "20.02.2026 09:00"])

    ws2 = wb.create_sheet("Projekte")
    ws2.append(["Nr.", "Kunde", "Projektname", "Aktueller Status", "Notiz", "Letzter Status Datum"])
    ws2.append([1001, "Muster GmbH", "PV Dach", "In Durchfuehrung", "Montage gestartet", "21.02.2026 07:30"])
    ws2.append([None, "Ohne Nummer GmbH", "Serviceeinsatz", "In Durchfuehrung", "Techniker vor Ort", "21.02.2026 10:15"])

    wb.save(file_path)

    db = SessionLocal()
    try:
        stats = import_projects_from_excel(db, str(file_path), source_label=file_path.name)
        assert stats.processed_rows == 4
        assert stats.duplicates_skipped == 2
        assert stats.created == 2
        assert stats.temporary_numbers == 1

        numbered = db.scalars(select(Project).where(Project.project_number == "1001")).first()
        assert numbered is not None
        assert numbered.status == "In Durchfuehrung"
        assert numbered.last_state == "Montage gestartet"
        assert numbered.last_status_at == datetime(2026, 2, 21, 7, 30)

        temporary = db.scalars(select(Project).where(Project.customer_name == "Ohne Nummer GmbH")).first()
        assert temporary is not None
        assert temporary.project_number.startswith("T")
        assert temporary.status == "In Durchfuehrung"
        assert temporary.last_state == "Techniker vor Ort"
        assert temporary.last_status_at == datetime(2026, 2, 21, 10, 15)

        second = import_projects_from_excel(db, str(file_path), source_label=file_path.name)
        assert second.created == 0
        assert second.updated == 2
        assert second.temporary_numbers == 0
        assert db.scalar(select(func.count(Project.id))) == 2
    finally:
        db.close()


def test_import_projects_from_excel_skips_rows_without_project_identity(tmp_path: Path):
    file_path = tmp_path / "projects_skip_empty_identity.xlsx"
    _write_workbook(
        file_path,
        [
            ["Nr.", "Kunde", "Projektname", "Aktueller Status"],
            [None, None, None, "Anfrage erhalten"],
            [501, "SMPL", "Gueltiges Projekt", "Angebot erstellen"],
        ],
    )

    db = SessionLocal()
    try:
        stats = import_projects_from_excel(db, str(file_path), source_label=file_path.name)
        assert stats.created == 1
        assert stats.temporary_numbers == 0
        project = db.scalars(select(Project).where(Project.project_number == "501")).first()
        assert project is not None
        assert project.name == "Gueltiges Projekt"
    finally:
        db.close()


def test_import_projects_from_csv_preserves_columns_and_generates_temp_numbers(tmp_path: Path):
    file_path = tmp_path / "projects.csv"
    _write_csv(
        file_path,
        [
            ["project_number", "name", "customer_name", "Aktueller Status", "Notiz", "Neue Spalte"],
            ["6001", "PV Nord", "CSV Kunde", "Angebot erstellen", "Initial", "A"],
            ["", "Service Einsatz", "CSV Kunde 2", "In Durchführung", "Temp Number", "B"],
        ],
    )

    db = SessionLocal()
    try:
        stats = import_projects_from_csv(db, str(file_path), source_label=file_path.name)
        assert stats.processed_rows == 2
        assert stats.created == 2
        assert stats.updated == 0
        assert stats.temporary_numbers == 1

        numbered = db.scalars(select(Project).where(Project.project_number == "6001")).first()
        assert numbered is not None
        assert numbered.name == "PV Nord"
        assert numbered.status == "Angebot erstellen"
        assert numbered.last_state == "Initial"
        assert numbered.extra_attributes["Neue Spalte"] == "A"

        temporary = db.scalars(select(Project).where(Project.customer_name == "CSV Kunde 2")).first()
        assert temporary is not None
        assert temporary.project_number.startswith("T")
        assert temporary.extra_attributes["Neue Spalte"] == "B"
    finally:
        db.close()
